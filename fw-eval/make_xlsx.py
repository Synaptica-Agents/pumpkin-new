# -*- coding: utf-8 -*-
"""Baut Frameworks-Bewertungstest.xlsx aus results.json (Vorbild: ms-eval).
Optionales Argument: Ausgabedateiname (Default Frameworks-Bewertungstest.xlsx)."""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(BASE)
rows = json.load(open(os.path.join(BASE, "results.json"), encoding="utf-8"))

wb = Workbook()
ws = wb.active
ws.title = "Bewertungen"

headers = ["Schwierigkeit", "Kategorie", "Case", "Musterlösung", "Input (an die KI)",
           "Erwartete Qualität", "KI-Feedback", "Score (0-100)",
           "Framework-Wahl /25", "Struktur & MECE /30", "Vollständigkeit /25",
           "Priorisierung /20", "Flagged"]
ws.append(headers)

hfill = PatternFill("solid", fgColor="1F2937")
hfont = Font(bold=True, color="FFFFFF")
for c in ws[1]:
    c.fill = hfill; c.font = hfont
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

tier_fill = {"perfekt": "DCFCE7", "normal": "FEF9C3", "schlecht": "FEE2E2"}
thin = Side(style="thin", color="E5E7EB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for r in rows:
    ws.append([r["difficulty"], r["category"], r["prompt"], r["muster"], r["input"],
               r["tier"], r["feedback"], r["score"], r["s_framework"], r["s_mece"],
               r["s_complete"], r["s_prio"], "ja" if r["flagged"] else ""])
    for c in ws[ws.max_row]:
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        fill = tier_fill.get(r["tier"])
        if fill: c.fill = PatternFill("solid", fgColor=fill)

widths = [12, 12, 34, 40, 44, 13, 46, 10, 12, 12, 12, 12, 8]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = "A2"

# --- Zusammenfassung ---
ws2 = wb.create_sheet("Zusammenfassung")

def num(x):
    try: return float(x)
    except (TypeError, ValueError): return None

def add_summary(title, subset):
    ws2.append([title])
    ws2[ws2.max_row][0].font = Font(bold=True, size=12)
    ws2.append(["Qualität", "n", "Ø Score", "Ø Framework/25", "Ø MECE/30",
                "Ø Vollst./25", "Ø Prio/20", "min", "max"])
    for c in ws2[ws2.max_row]:
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for tier in ("perfekt", "normal", "schlecht"):
        sub = [r for r in subset if r["tier"] == tier and num(r["score"]) is not None]
        if not sub:
            ws2.append([tier, 0]); continue
        sc = [num(r["score"]) for r in sub]
        def avg(key):
            v = [num(r[key]) for r in sub if num(r[key]) is not None]
            return round(sum(v) / len(v), 1) if v else ""
        ws2.append([tier, len(sub), round(sum(sc) / len(sc), 1),
                    avg("s_framework"), avg("s_mece"), avg("s_complete"),
                    avg("s_prio"), min(sc), max(sc)])
        fill = tier_fill.get(tier)
        for c in ws2[ws2.max_row]:
            if fill: c.fill = PatternFill("solid", fgColor=fill)
    ws2.append([])

add_summary("Gesamt (30 Cases × 3 Eingaben)", rows)
add_summary("Nur Mittel", [r for r in rows if r["difficulty"] == "medium"])
add_summary("Nur Schwer", [r for r in rows if r["difficulty"] == "hard"])

for i, w in enumerate([14, 6, 10, 14, 12, 12, 10, 8, 8], 1):
    ws2.column_dimensions[chr(64 + i)].width = w

out = os.path.join(ROOT, sys.argv[1] if len(sys.argv) > 1 else "Frameworks-Bewertungstest.xlsx")
wb.save(out)
print("geschrieben:", out)
