# -*- coding: utf-8 -*-
"""Baut Diagramme-Bewertungstest.xlsx aus results.json (Vorbild: fw-eval).
Optionales Argument: Ausgabedateiname (Default Diagramme-Bewertungstest.xlsx)."""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(BASE)
rows = json.load(open(os.path.join(BASE, "results.json"), encoding="utf-8"))

wb = Workbook()
ws = wb.active
ws.title = "Bewertungen"

headers = ["Schwierigkeit", "Exhibit-Typ", "Case", "Referenzlösung", "Input (an die KI)",
           "Erwartete Qualität", "KI-Feedback", "Score (0-100)",
           "Daten-Ablesung /25", "Trend-Analyse /25", "Business-Implikationen /25",
           "Analysetiefe /15", "Kommunikation /10", "Flagged"]
ws.append(headers)

hfill = PatternFill("solid", fgColor="1F2937")
hfont = Font(bold=True, color="FFFFFF")
for c in ws[1]:
    c.fill = hfill; c.font = hfont
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

tier_fill = {"perfekt": "DCFCE7", "normal": "FEF9C3", "schlecht": "FEE2E2"}
thin = Side(style="thin", color="E5E7EB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for r in rows:
    ws.append([r["difficulty"], r["category"], r["prompt"], r["muster"], r["input"],
               r["tier"], r["feedback"], r["score"], r["s_data"], r["s_trend"],
               r["s_biz"], r["s_depth"], r["s_comm"], "ja" if r["flagged"] else ""])
    for c in ws[ws.max_row]:
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        fill = tier_fill.get(r["tier"])
        if fill: c.fill = PatternFill("solid", fgColor=fill)

widths = [12, 12, 34, 40, 44, 13, 46, 10, 12, 12, 14, 12, 12, 8]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = "A2"

# --- Zusammenfassung ---
ws2 = wb.create_sheet("Zusammenfassung")

def num(x):
    try: return float(x)
    except (TypeError, ValueError): return None

def add_summary(title, subset):
    ws2.append([title])
    ws2[ws2.max_row][0].font = Font(bold=True, size=12)
    ws2.append(["Qualität", "n", "Ø Score", "Ø Daten/25", "Ø Trend/25",
                "Ø Business/25", "Ø Tiefe/15", "Ø Komm./10", "min", "max"])
    for c in ws2[ws2.max_row]:
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for tier in ("perfekt", "normal", "schlecht"):
        sub = [r for r in subset if r["tier"] == tier and num(r["score"]) is not None]
        if not sub:
            ws2.append([tier, 0]); continue
        sc = [num(r["score"]) for r in sub]
        def avg(key):
            v = [num(r[key]) for r in sub if num(r[key]) is not None]
            return round(sum(v) / len(v), 1) if v else ""
        ws2.append([tier, len(sub), round(sum(sc) / len(sc), 1),
                    avg("s_data"), avg("s_trend"), avg("s_biz"),
                    avg("s_depth"), avg("s_comm"), min(sc), max(sc)])
        fill = tier_fill.get(tier)
        for c in ws2[ws2.max_row]:
            if fill: c.fill = PatternFill("solid", fgColor=fill)
    ws2.append([])

add_summary("Gesamt (30 Cases × 3 Eingaben)", rows)
add_summary("Nur Einsteiger (easy)", [r for r in rows if r["difficulty"] == "easy"])
add_summary("Nur Fortgeschritten (hard)", [r for r in rows if r["difficulty"] == "hard"])

for i, w in enumerate([14, 6, 10, 12, 12, 14, 10, 10, 8, 8], 1):
    ws2.column_dimensions[chr(64 + i)].width = w

out = os.path.join(ROOT, sys.argv[1] if len(sys.argv) > 1 else "Diagramme-Bewertungstest.xlsx")
wb.save(out)
print("geschrieben:", out)
